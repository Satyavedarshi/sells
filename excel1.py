
import openpyxl, os

filepath = '/Users/saramise/Downloads/dates2.xlsx'
copyfile = '/Users/saramise/Downloads/dates2.xlsx'

workbook = openpyxl.load_workbook(filepath)
sheet1 = workbook.active
print(sheet1.max_row, sheet1.max_column)

workbook1 = openpyxl.load_workbook(copyfile)
sheet2 = workbook1.active

for rows in range(1, sheet1.max_row):
    for cols in range(1, sheet1.max_column):
        print(sheet1.cell(rows,cols).value)
        sheet2.cell(rows,cols).value = 'welcome_new'

workbook1.save(copyfile)

